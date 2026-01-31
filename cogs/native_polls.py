import discord
from discord.ext import commands
from datetime import datetime
import io
from utils import is_admin_or_whitelisted, t


class NativePollSystem(commands.Cog):
    """Экспорт данных опросов Discord"""

    def __init__(self, bot):
        self.bot = bot
        self.db = bot.db

    @commands.command(name='gb_poll_export_detailed')
    @is_admin_or_whitelisted()
    async def gb_poll_export_detailed(self, ctx, message_id_or_link: str, days: int = 7):
        """
        Экспорт опроса напрямую из Discord с детальной статистикой активности.
        Работает только для ЗАВЕРШЁННЫХ опросов.

        Формат:
        !gb_poll_export_detailed <message_id> [7/14/30]
        !gb_poll_export_detailed <ссылка на сообщение> [7/14/30]
        """
        await ctx.message.delete()
        guild_id = ctx.guild.id

        if days not in [7, 14, 30]:
            await ctx.send(t('poll_invalid_period', guild_id=guild_id), delete_after=10)
            return

        # Парсим ID из ссылки
        msg_id = None
        channel_id = None

        if 'discord.com/channels/' in message_id_or_link or 'discordapp.com/channels/' in message_id_or_link:
            try:
                parts = message_id_or_link.rstrip('/').split('/')
                msg_id = int(parts[-1])
                channel_id = int(parts[-2])
            except (ValueError, IndexError):
                await ctx.send(t('poll_invalid_link', guild_id=guild_id), delete_after=10)
                return
        else:
            try:
                msg_id = int(message_id_or_link)
            except ValueError:
                await ctx.send(t('poll_invalid_id', guild_id=guild_id), delete_after=10)
                return

        # Получаем канал
        if channel_id:
            channel = self.bot.get_channel(channel_id)
        else:
            channel = ctx.channel

        if not channel:
            await ctx.send(t('poll_channel_not_found', guild_id=guild_id), delete_after=10)
            return

        try:
            message = await channel.fetch_message(msg_id)
        except discord.NotFound:
            await ctx.send(t('poll_message_not_found', guild_id=guild_id), delete_after=10)
            return
        except discord.Forbidden:
            await ctx.send(t('poll_no_access', guild_id=guild_id), delete_after=10)
            return

        poll = message.poll
        if not poll:
            await ctx.send(t('poll_not_a_poll', guild_id=guild_id), delete_after=10)
            return

        if not poll.is_finalized():
            await ctx.send(t('poll_not_finalized', guild_id=guild_id), delete_after=15)
            return

        # Собираем голоса напрямую из Discord API
        status_msg = await ctx.send(t('poll_loading_votes', guild_id=guild_id))

        votes_by_answer = {}  # {answer_id: [members]}
        all_voters = set()

        for answer in poll.answers:
            try:
                voters = []
                async for voter in answer.voters():
                    voters.append(voter)
                    all_voters.add(voter.id)
                votes_by_answer[answer.id] = voters
            except Exception as e:
                print(f"Error fetching voters for answer {answer.id}: {e}")
                votes_by_answer[answer.id] = []

        total_votes = sum(len(v) for v in votes_by_answer.values())

        if total_votes == 0:
            await status_msg.delete()
            await ctx.send(t('poll_no_votes', guild_id=guild_id), delete_after=10)
            return

        # Получаем статистику активности для всех проголосовавших
        await status_msg.edit(content=t('poll_loading_stats', guild_id=guild_id))

        user_stats = {}
        for user_id in all_voters:
            stats = self.db.get_user_stats(ctx.guild.id, user_id, days)
            if stats:
                user_stats[user_id] = {
                    'messages': stats['period_messages'],
                    'voice_time': stats['period_voice_time']
                }
            else:
                user_stats[user_id] = {'messages': 0, 'voice_time': 0}

        # Сортируем по времени в войсе
        sorted_votes_by_answer = {}
        for answer_id, voters in votes_by_answer.items():
            sorted_voters = sorted(
                voters,
                key=lambda m: user_stats.get(m.id, {}).get('voice_time', 0),
                reverse=True
            )
            sorted_votes_by_answer[answer_id] = sorted_voters

        await status_msg.edit(content=t('poll_creating_file', guild_id=guild_id))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Poll Results Detailed"

            # Заголовок
            ws['A1'] = 'Discord Poll Export (Fetched from Discord)'
            ws['A1'].font = Font(bold=True, size=14)

            ws['A2'] = 'Message ID:'
            ws['B2'] = msg_id
            ws['A3'] = 'Question:'
            ws['B3'] = str(poll.question)
            ws['A4'] = 'Period:'
            ws['B4'] = f'{days} days'

            ws['A6'] = 'Total Votes:'
            ws['B6'] = total_votes

            # Статистика по вариантам
            ws['A8'] = 'Option'
            ws['B8'] = 'Votes'
            ws['C8'] = 'Percentage'
            ws['A8'].font = Font(bold=True)
            ws['B8'].font = Font(bold=True)
            ws['C8'].font = Font(bold=True)

            row = 9
            for answer in poll.answers:
                count = len(votes_by_answer.get(answer.id, []))
                percentage = (count / total_votes * 100) if total_votes > 0 else 0
                ws[f'A{row}'] = answer.text
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f"{percentage:.1f}%"
                row += 1

            # Пустая строка
            row += 1

            # Заголовки вариантов
            start_col = 1  # Колонка B
            header_row = row

            yellow_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for i, answer in enumerate(poll.answers):
                col_letter = chr(ord('A') + start_col + i)
                cell = ws[f'{col_letter}{header_row}']

                cell.value = answer.text
                cell.fill = yellow_fill
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

                ws.column_dimensions[col_letter].width = 35

            # Заполняем данные с статистикой
            max_votes = max([len(sorted_votes_by_answer.get(answer.id, [])) for answer in poll.answers], default=0)

            for row_index in range(max_votes):
                data_row = header_row + 1 + row_index

                for i, answer in enumerate(poll.answers):
                    voters = sorted_votes_by_answer.get(answer.id, [])
                    col_letter = chr(ord('A') + start_col + i)

                    if row_index < len(voters):
                        member = voters[row_index]
                        user_id = member.id

                        cell = ws[f'{col_letter}{data_row}']

                        username = member.display_name
                        profile_link = f"https://discord.com/users/{user_id}"

                        stats = user_stats.get(user_id, {'messages': 0, 'voice_time': 0})
                        messages = stats['messages']
                        voice_hours = int(stats['voice_time'] // 3600)
                        voice_minutes = int((stats['voice_time'] % 3600) // 60)

                        cell_value = f"{username} | {messages} msg | {voice_hours}h {voice_minutes}m"

                        cell.value = cell_value
                        cell.hyperlink = profile_link
                        cell.font = Font(color='0563C1', underline='single')
                        cell.alignment = Alignment(vertical='center')
                        cell.border = thin_border

            # Сохраняем
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            file = discord.File(
                excel_file,
                filename=f'poll_{msg_id}_fetched_{days}d.xlsx'
            )

            await status_msg.delete()
            await ctx.send(t('poll_export_success', guild_id=guild_id, days=days, votes=total_votes), file=file)

        except ImportError:
            await status_msg.delete()
            await ctx.send(t('poll_openpyxl_required', guild_id=guild_id), delete_after=10)
        except Exception as e:
            await status_msg.delete()
            await ctx.send(t('poll_error', guild_id=guild_id, error=str(e)), delete_after=10)
            import traceback
            traceback.print_exc()


async def setup(bot):
    await bot.add_cog(NativePollSystem(bot))
